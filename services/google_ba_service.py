# services/google_ba_service.py - Google Service untuk Berita Acara
import os
import json
import base64
import logging
import io
import tempfile
import shutil
from datetime import datetime, timedelta
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from googleapiclient.errors import HttpError
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from oauth_token_manager import get_access_token

logger = logging.getLogger(__name__)

# Scopes untuk Google API
SCOPES = [
    'https://www.googleapis.com/auth/drive',
    'https://www.googleapis.com/auth/spreadsheets'
]

class GoogleBAService:
    def __init__(self, template_folder_id, result_folder_id):
        self.template_folder_id = template_folder_id
        self.result_folder_id = result_folder_id
        self.oauth_client_config = os.environ.get('GOOGLE_OAUTH_CLIENT_CONFIG')
        
        # Services
        self.service_drive = None
        self.service_sheets = None
        self.credentials = None
        
        # Token management
        self.token_file = 'token.json'
        
        # Validate environment
        self._validate_environment()

    def _validate_environment(self):
        """Validate required environment variables"""
        if not self.oauth_client_config:
            raise ValueError("GOOGLE_OAUTH_CLIENT_CONFIG environment variable is required")
        
        if not self.template_folder_id:
            raise ValueError("TEMPLATE_FOLDER_ID environment variable is required")
            
        if not self.result_folder_id:
            raise ValueError("RESULT_FOLDER_ID environment variable is required")
        
        logger.info("✅ Environment variables validated")

    def authenticate(self):
        """Hanya menggunakan token dari oauth_token_manager.py"""
        try:
            # Gunakan token dari oauth_token_manager
            if not self.ensure_valid_token():
                return False
            
            # Build services dengan credentials yang sudah ada
            self.service_drive = build('drive', 'v3', credentials=self.credentials)
            self.service_sheets = build('sheets', 'v4', credentials=self.credentials)
            
            logger.info("✅ Google APIs authenticated successfully with OAuth")
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to authenticate Google APIs: {e}")
            return False

    def ensure_valid_token(self):
        """Ensure we have a valid token, refresh if needed"""
        token = get_access_token()
        if token:
            # Update credentials object dengan token baru
            if self.credentials:
                self.credentials.token = token
                return True
            else:
                # Buat credentials baru jika belum ada
                from google.oauth2.credentials import Credentials
                self.credentials = Credentials(token=token)
                
                # Rebuild services dengan credentials baru
                self.service_drive = build('drive', 'v3', credentials=self.credentials)
                self.service_sheets = build('sheets', 'v4', credentials=self.credentials)
                return True
        
        logger.error("Failed to get valid access token")
        return False

    # Semua method lainnya tetap sama, tapi tambahkan ensure_valid_token() di awal setiap method yang menggunakan API

    def find_excel_template(self):
        """Find Excel template file in template folder"""
        try:
            if not self.ensure_valid_token():
                return None
                
            # Search for Excel files in template folder
            query = f"'{self.template_folder_id}' in parents and (mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or mimeType='application/vnd.ms-excel')"
            
            results = self.service_drive.files().list(
                q=query,
                fields="files(id, name, mimeType)",
                supportsAllDrives=True
            ).execute()
            
            files = results.get('files', [])
            
            if not files:
                logger.error("❌ No Excel template found in template folder")
                return None
            
            # Use first Excel file found
            template_file = files[0]
            logger.info(f"✅ Template found: {template_file['name']} (ID: {template_file['id']})")
            
            return template_file
            
        except Exception as e:
            logger.error(f"❌ Error finding template: {e}")
            return None

    # Tambahkan ensure_valid_token() di semua method yang menggunakan Google API
    # Contoh untuk method download_excel_template:
    def download_excel_template(self, file_id):
        """Download Excel template to temporary file"""
        try:
            if not self.ensure_valid_token():
                return None
                
            # Download file
            request = self.service_drive.files().get_media(fileId=file_id)
            
            # Create temporary file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            
            downloader = MediaIoBaseDownload(temp_file, request)
            done = False
            
            while done is False:
                status, done = downloader.next_chunk()
                
            temp_file.close()
            
            logger.info(f"✅ Template downloaded to: {temp_file.name}")
            return temp_file.name
            
        except Exception as e:
            logger.error(f"❌ Error downloading template: {e}")
            return None

    # Lakukan hal yang sama untuk semua method yang menggunakan Google API:
    # convert_excel_to_pdf, upload_pdf_result, create_evidence_folder, 
    # upload_photo_evidence, get_drive_info, dll.

    def fill_excel_template(self, template_path, form_data, ba_config, form_type='wifi'):
        """Fill Excel template with form data, including signature images that fit properly in cells"""
        temp_files_to_cleanup = []
        
        try:
            logger.info("📋 Filling Excel template with form data and signatures...")
            
            # Load workbook
            workbook = openpyxl.load_workbook(template_path)
            worksheet = workbook.active
            
            # Prepare Excel data using ba_config
            excel_data = ba_config.prepare_excel_data(form_data, form_type)
            
            # Fill data into cells
            for coordinate, value in excel_data.items():
                try:
                    worksheet[coordinate] = value
                    logger.debug(f"📝 Filled {coordinate}: {value}")
                except Exception as e:
                    logger.warning(f"⚠️ Could not fill cell {coordinate}: {e}")
            
            # Handle signature images dengan ukuran yang pas di dalam sel
            tanda_tangan_section = form_data.get('tanda_tangan', {})
            signature_files = []
            
            if tanda_tangan_section:
                for field_name, image_path in tanda_tangan_section.items():
                    if image_path and image_path.startswith('SIGNATURE_IMAGE:'):
                        actual_path = image_path.replace('SIGNATURE_IMAGE:', '')
                        coordinate = ba_config.get_excel_coordinates('tanda_tangan', field_name, form_type)
                        
                        if coordinate and os.path.exists(actual_path):
                            try:
                                # PERBAIKAN: Fit gambar ke dalam sel dengan ukuran yang tepat
                                cell = worksheet[coordinate]
                                
                                # Dapatkan ukuran sel dalam pixels menggunakan helper method
                                cell_width_px, cell_height_px = self._calculate_cell_dimensions(worksheet, coordinate)
                                
                                # Load image
                                img = XLImage(actual_path)
                                
                                # BAGIAN PENGATURAN SKALA - CUSTOM DISINI
                                # =============================================
                                # Margin dari tepi sel (dalam pixels)
                                margin = 0.5  # CUSTOM: ubah nilai ini untuk mengatur jarak dari tepi sel
                                
                                # Faktor skala custom (0.1 = 10%, 1.0 = 100%, 1.5 = 150%)
                                custom_scale_factor = 1.4  # CUSTOM: ubah nilai ini untuk memperbesar/memperkecil gambar secara keseluruhan
                                
                                # Ukuran maksimum yang tersedia dalam sel (dikurangi margin)
                                available_width = max(cell_width_px - (margin * 2), 50)  # minimal 50px
                                available_height = max(cell_height_px - (margin * 2), 30)  # minimal 30px
                                
                                # Terapkan custom scale factor ke ukuran yang tersedia
                                target_width = available_width * custom_scale_factor
                                target_height = available_height * custom_scale_factor
                                
                                # Dapatkan dimensi asli gambar signature
                                # Untuk gambar dengan dimensi 365x380 px dan 301 DPI
                                original_img_width = img.width
                                original_img_height = img.height
                                
                                logger.info(f"📏 Original image dimensions: {original_img_width}x{original_img_height}px")
                                logger.info(f"📐 Cell dimensions: {cell_width_px}x{cell_height_px}px")
                                logger.info(f"🎯 Target dimensions: {target_width}x{target_height}px")
                                
                                # Hitung rasio skala untuk mempertahankan aspect ratio
                                width_ratio = target_width / original_img_width
                                height_ratio = target_height / original_img_height
                                
                                # Ambil rasio yang lebih kecil agar gambar tidak keluar dari area target
                                scale_ratio = min(width_ratio, height_ratio)
                                
                                # Hitung ukuran final
                                final_width = original_img_width * scale_ratio
                                final_height = original_img_height * scale_ratio
                                
                                # FINE-TUNING: Manual override untuk ukuran tertentu (opsional)
                                # Uncomment dan sesuaikan jika ingin ukuran fixed
                                # final_width = 120   # CUSTOM: ukuran lebar fixed dalam pixels
                                # final_height = 60   # CUSTOM: ukuran tinggi fixed dalam pixels
                                
                                # Set ukuran gambar
                                img.width = int(final_width)
                                img.height = int(final_height)
                                # =============================================
                                # AKHIR BAGIAN PENGATURAN SKALA
                                
                                # Set anchor di koordinat sel
                                img.anchor = coordinate
                                
                                # Add image ke worksheet
                                worksheet.add_image(img)
                                signature_files.append(actual_path)
                                
                                logger.info(f"✅ Added signature image at {coordinate} "
                                        f"(final size: {int(final_width)}x{int(final_height)}px, "
                                        f"scale ratio: {scale_ratio:.3f})")
                                        
                            except Exception as e:
                                logger.error(f"❌ Error adding signature image: {e}")
                                # Fallback: add image dengan ukuran default yang disesuaikan
                                try:
                                    img = XLImage(actual_path)
                                    # Fallback size yang disesuaikan dengan signature 365x380
                                    fallback_width = 100  # CUSTOM: ubah fallback width
                                    fallback_height = 52   # CUSTOM: ubah fallback height (mempertahankan rasio 365:380)
                                    
                                    img.width = fallback_width
                                    img.height = fallback_height
                                    img.anchor = coordinate
                                    worksheet.add_image(img)
                                    signature_files.append(actual_path)
                                    logger.warning(f"⚠️ Added signature with fallback size at {coordinate} "
                                                f"({fallback_width}x{fallback_height}px)")
                                except Exception as fallback_error:
                                    logger.error(f"❌ Fallback failed: {fallback_error}")
            
            # Save filled template to temporary file
            temp_dir = tempfile.gettempdir()
            filled_filename = f"filled_ba_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filled_path = os.path.join(temp_dir, filled_filename)
            
            workbook.save(filled_path)
            logger.info(f"✅ Filled template saved: {filled_path}")
            
            # Clean up signature files after adding to Excel
            for sig_file in signature_files:
                try:
                    if os.path.exists(sig_file):
                        os.remove(sig_file)
                        logger.info(f"🗑️ Cleaned up signature file: {sig_file}")
                except Exception as e:
                    logger.warning(f"⚠️ Could not clean up signature file {sig_file}: {e}")
            
            return filled_path
            
        except Exception as e:
            logger.error(f"❌ Error filling Excel template: {e}")
            
            # Clean up any created files on error
            for file_path in temp_files_to_cleanup:
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                except:
                    pass
                
    def delete_folder(self, folder_id):
        """Delete a Google Drive folder"""
        try:
            if not self.ensure_valid_token():
                return False
                
            self.service_drive.files().delete(fileId=folder_id).execute()
            logger.info(f"✅ Folder deleted: {folder_id}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Error deleting folder {folder_id}: {e}")
            return False

    # Helper method untuk menghitung ukuran sel yang lebih akurat
    def _calculate_cell_dimensions(self, worksheet, coordinate):
        """Calculate cell dimensions in pixels more accurately"""
        try:
            from openpyxl.utils import coordinate_from_string, column_index_from_string, get_column_letter
            
            col_letter, row_num = coordinate_from_string(coordinate)
            col_index = column_index_from_string(col_letter)
            
            # Get column width (dalam character units)
            col_dim = worksheet.column_dimensions.get(col_letter)
            col_width_chars = col_dim.width if col_dim and col_dim.width else 8.43  # Excel default
            
            # Get row height (dalam point units)  
            row_dim = worksheet.row_dimensions.get(row_num)
            row_height_points = row_dim.height if row_dim and row_dim.height else 15  # Excel default
            
            # Konversi ke pixels dengan faktor konversi yang akurat untuk DPI 301
            # Disesuaikan dengan resolusi signature image 301 DPI
            # 1 character width ≈ 7.5 pixels (untuk font default Excel)
            # 1 point ≈ 1.33 pixels
            cell_width_px = col_width_chars * 7.5
            cell_height_px = row_height_points * 1.33
            
            return int(cell_width_px), int(cell_height_px)
            
        except Exception as e:
            logger.warning(f"Could not calculate cell dimensions: {e}")
            # Fallback dimensions yang disesuaikan dengan signature 365x380
            return 120, 80  # fallback dimensions yang lebih proporsional

    def upload_excel_result(self, excel_path, filename, folder_id=None):
        """Upload final Excel to specific folder (default: result folder)"""
        try:
            if not self.ensure_valid_token():
                return None

            logger.info(f"📤 Uploading Excel result: {filename}")
            
            # Use provided folder or default result folder
            parents = [folder_id] if folder_id else [self.result_folder_id]
            
            file_metadata = {
                'name': f"{filename}.xlsx",
                'parents': parents
            }
            
            media = MediaFileUpload(excel_path, 
                                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                                resumable=True)
            
            uploaded_file = self.service_drive.files().create(
                body=file_metadata,
                media_body=media,
                supportsAllDrives=True
            ).execute()
            
            file_id = uploaded_file.get('id')
            
            # Generate shareable link
            link = f"https://drive.google.com/file/d/{file_id}/view"
            
            logger.info(f"✅ Excel uploaded successfully: {file_id}")
            return link
            
        except Exception as e:
            logger.error(f"❌ Error uploading Excel: {e}")
            return None

    def create_folder_structure(self, base_folder_id, folder_name):
        """Create organized folder structure for reports"""
        try:
            if not self.ensure_valid_token():
                return None, None, None

            # Create main report folder
            report_folder_metadata = {
                'name': folder_name,
                'mimeType': 'application/vnd.google-apps.folder',
                'parents': [base_folder_id]
            }
            
            report_folder = self.service_drive.files().create(
                body=report_folder_metadata,
                supportsAllDrives=True
            ).execute()
            
            report_folder_id = report_folder.get('id')
            
            # Create subfolders
            evidence_folder_metadata = {
                'name': 'Evidence',
                'mimeType': 'application/vnd.google-apps.folder',
                'parents': [report_folder_id]
            }
            
            evidence_folder = self.service_drive.files().create(
                body=evidence_folder_metadata,
                supportsAllDrives=True
            ).execute()
            
            evidence_folder_id = evidence_folder.get('id')
            
            # Create BA Form folder
            ba_form_folder_metadata = {
                'name': 'Form BA',
                'mimeType': 'application/vnd.google-apps.folder',
                'parents': [report_folder_id]
            }
            
            ba_form_folder = self.service_drive.files().create(
                body=ba_form_folder_metadata,
                supportsAllDrives=True
            ).execute()
            
            ba_form_folder_id = ba_form_folder.get('id')
            
            logger.info(f"📁 Folder structure created: {folder_name}")
            return report_folder_id, evidence_folder_id, ba_form_folder_id
                
        except Exception as e:
            logger.error(f"❌ Error creating folder structure: {e}")
            return None, None, None

    async def process_excel_only(self, form_data, filename, ba_config, form_type='wifi'):
        """Complete process with organized folder structure"""
        temp_files = []
        
        try:
            logger.info("🚀 Starting Excel processing with organized folders...")
            
            # Step 1: Find Excel template
            template_file = self.find_excel_template()
            if not template_file:
                return False, "Template Excel tidak ditemukan di folder template"
            
            # Step 2: Download template
            template_path = self.download_excel_template(template_file['id'])
            if not template_path:
                return False, "Gagal download template Excel"
            
            temp_files.append(template_path)
            
            # Step 3: Fill template with data
            filled_path = self.fill_excel_template(template_path, form_data, ba_config, form_type)
            if not filled_path:
                return False, "Gagal mengisi template Excel"
            
            temp_files.append(filled_path)
            
            # Step 4: Create organized folder structure
            folder_name = filename  # Use the generated filename as folder name
            report_folder_id, evidence_folder_id, ba_form_folder_id = self.create_folder_structure(
                self.result_folder_id, folder_name
            )
            
            if not all([report_folder_id, evidence_folder_id, ba_form_folder_id]):
                return False, "Gagal membuat struktur folder"
            
            # Step 5: Upload Excel to Form BA folder
            result_link = self.upload_excel_result(filled_path, filename, ba_form_folder_id)
            if not result_link:
                return False, "Gagal upload Excel result"
            
            # Step 6: Store folder IDs for evidence uploads
            result_info = {
                'excel_link': result_link,
                'report_folder_id': report_folder_id,
                'evidence_folder_id': evidence_folder_id,
                'ba_form_folder_id': ba_form_folder_id,
                'report_folder_link': self.get_folder_link(report_folder_id),
                'evidence_folder_link': self.get_folder_link(evidence_folder_id),
                'ba_form_folder_link': self.get_folder_link(ba_form_folder_id)
            }
            
            # Cleanup temp files
            self.cleanup_temp_files(*temp_files)
            
            logger.info("✅ Excel processing with organized folders completed successfully!")
            return True, result_info
            
        except Exception as e:
            logger.error(f"❌ Error in process_excel_only: {e}")
            
            # Cleanup temp files on error
            self.cleanup_temp_files(*temp_files)
            
            return False, f"Terjadi kesalahan: {str(e)}"


    def create_evidence_folder(self, folder_name):
        """Create folder for evidence photos"""
        try:
            
            if not self.ensure_valid_token():
                return None
            
            folder_metadata = {
                'name': folder_name,
                'mimeType': 'application/vnd.google-apps.folder',
                'parents': [self.result_folder_id]
            }
            
            folder = self.service_drive.files().create(
                body=folder_metadata,
                supportsAllDrives=True
            ).execute()
            
            folder_id = folder.get('id')
            logger.info(f"📁 Evidence folder created: {folder_name} (ID: {folder_id})")
            
            return folder_id
            
        except Exception as e:
            logger.error(f"❌ Error creating evidence folder: {e}")
            return None

    def upload_photo_evidence(self, photo_path, filename, folder_id):
        """Upload photo evidence to specific folder"""
        try:
            
            if not self.ensure_valid_token():
                return None
            
            file_metadata = {
                'name': filename,
                'parents': [folder_id]
            }
            
            media = MediaFileUpload(
                photo_path,
                mimetype='image/jpeg',
                resumable=True
            )
            
            uploaded_file = self.service_drive.files().create(
                body=file_metadata,
                media_body=media,
                supportsAllDrives=True
            ).execute()
            
            file_id = uploaded_file.get('id')
            logger.info(f"📷 Photo uploaded: {filename} -> {file_id}")
            
            return file_id
            
        except Exception as e:
            logger.error(f"❌ Error uploading photo: {e}")
            return None

    def get_folder_link(self, folder_id):
        """Get shareable link for Google Drive folder"""
        return f"https://drive.google.com/drive/folders/{folder_id}"

    def get_file_link(self, file_id):
        """Get shareable link for Google Drive file"""
        return f"https://drive.google.com/file/d/{file_id}/view"

    def cleanup_temp_files(self, *file_paths):
        """Clean up temporary files with retry logic"""
        import time
        for file_path in file_paths:
            try:
                if file_path and os.path.exists(file_path):
                    # Try multiple times with delay
                    for attempt in range(3):
                        try:
                            os.remove(file_path)
                            logger.debug(f"🗑️ Cleaned up: {file_path}")
                            break
                        except PermissionError:
                            if attempt < 2:  # Don't wait on last attempt
                                time.sleep(0.1)
                            continue
            except Exception as e:
                logger.warning(f"⚠️ Could not clean up {file_path}: {e}")


    def _has_photos_in_form(self, form_data):
        """Check if form data indicates photos will be uploaded"""
        # This is a placeholder - in practice, photos would be handled separately
        # For now, we'll create evidence folder if keterangan section mentions photos
        keterangan_section = form_data.get('keterangan', {})
        if keterangan_section:
            keterangan_text = keterangan_section.get('KETERANGAN TAMBAHAN', '').lower()
            return 'foto' in keterangan_text or 'photo' in keterangan_text or 'gambar' in keterangan_text
        
        return False

    def validate_template_structure(self, template_path, ba_config):
        """Validate that template has expected structure"""
        try:
            workbook = openpyxl.load_workbook(template_path)
            worksheet = workbook.active
            
            # Check if key coordinates exist and are accessible
            test_coordinates = ['O4', 'S4', 'O5', 'D4', 'D5']  # Sample coordinates
            
            valid_count = 0
            for coord in test_coordinates:
                try:
                    cell = worksheet[coord]
                    valid_count += 1
                except:
                    logger.warning(f"⚠️ Coordinate {coord} not accessible")
            
            if valid_count < len(test_coordinates) * 0.8:  # At least 80% should be valid
                logger.warning("⚠️ Template structure may not match expected coordinates")
                return False
            
            logger.info("✅ Template structure validation passed")
            return True
            
        except Exception as e:
            logger.error(f"❌ Error validating template: {e}")
            return False

    def get_drive_info(self):
        """Get information about Drive folders and template"""
        try:
            
            if not self.ensure_valid_token():
                return None
            
            info = {
                'template_folder': {},
                'result_folder': {},
                'template_file': None
            }
            
            # Get template folder info
            try:
                template_folder = self.service_drive.files().get(
                    fileId=self.template_folder_id,
                    supportsAllDrives=True
                ).execute()
                
                info['template_folder'] = {
                    'name': template_folder.get('name'),
                    'id': self.template_folder_id
                }
            except Exception as e:
                logger.error(f"❌ Error getting template folder info: {e}")
            
            # Get result folder info
            try:
                result_folder = self.service_drive.files().get(
                    fileId=self.result_folder_id,
                    supportsAllDrives=True
                ).execute()
                
                info['result_folder'] = {
                    'name': result_folder.get('name'),
                    'id': self.result_folder_id
                }
            except Exception as e:
                logger.error(f"❌ Error getting result folder info: {e}")
            
            # Find template file
            template_file = self.find_excel_template()
            if template_file:
                info['template_file'] = {
                    'name': template_file.get('name'),
                    'id': template_file.get('id'),
                    'mimeType': template_file.get('mimeType')
                }
            
            return info
            
        except Exception as e:
            logger.error(f"❌ Error getting drive info: {e}")
            return None

    def test_template_access(self):
        """Test if template can be accessed and processed"""
        try:
            if not self.ensure_valid_token():
                return None
            
            logger.info("🧪 Testing template access...")
            
            # Find template
            template_file = self.find_excel_template()
            if not template_file:
                return False, "Template file not found"
            
            # Try to download template
            template_path = self.download_excel_template(template_file['id'])
            if not template_path:
                return False, "Cannot download template"
            
            # Try to open with openpyxl
            try:
                workbook = openpyxl.load_workbook(template_path)
                worksheet = workbook.active
                logger.info("✅ Template can be opened with openpyxl")
            except Exception as e:
                return False, f"Cannot open template with openpyxl: {e}"
            
            # Test writing to a coordinate
            try:
                worksheet['A1'] = 'Test'
                test_path = template_path.replace('.xlsx', '_test.xlsx')
                workbook.save(test_path)
                logger.info("✅ Template can be modified and saved")
                
                # Cleanup test files
                self.cleanup_temp_files(template_path, test_path)
                
            except Exception as e:
                return False, f"Cannot modify template: {e}"
            
            return True, "Template access test passed"
            
        except Exception as e:
            logger.error(f"❌ Error in template access test: {e}")
            return False, f"Template access test failed: {e}"
